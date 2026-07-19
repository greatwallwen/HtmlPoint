from __future__ import annotations

import os
from pathlib import Path

import duckdb
import pytest
from openpyxl import Workbook

from course_helper.domain.common import SourceLocator
from course_helper.parsers import dataset_profiler as dataset_profiler_module
from course_helper.parsers.dataset_profiler import DatasetProfiler
from course_helper.source_roots import SourceRootRegistry


pytestmark = pytest.mark.filterwarnings(
    "ignore:datetime.datetime.utcnow.*:DeprecationWarning:openpyxl.*"
)


def write(path: Path, text: str) -> Path:
    path.write_text(text, encoding="utf-8")
    return path


def locator(relative_path: str) -> SourceLocator:
    return SourceLocator(root_id="fixture", relative_path=relative_path)


def profiler_for(root: Path) -> DatasetProfiler:
    return DatasetProfiler(SourceRootRegistry({"fixture": root}))


def reference_profiler() -> DatasetProfiler:
    configured_root = os.environ.get("COURSE_REFERENCE_ROOT")
    if not configured_root:
        pytest.skip("COURSE_REFERENCE_ROOT is required for reference_demo tests")
    return DatasetProfiler(
        SourceRootRegistry({"reference-demo": Path(configured_root)})
    )


@pytest.mark.parametrize("name", ["model.pth", "model.pt", "partial.tmp", "package.whl"])
def test_inventory_quarantines_non_dataset_payloads(tmp_path: Path, name: str) -> None:
    (tmp_path / name).write_bytes(b"not executable by the helper")

    item = profiler_for(tmp_path).inventory_directory(locator("."))[0]

    assert item.disposition == "quarantined"


def test_csv_profile_records_schema_grain_and_bounded_sample(tmp_path: Path) -> None:
    write(tmp_path / "sales.csv", "order_id,customer,amount\n1,A,10\n2,B,20\n")

    profile = profiler_for(tmp_path).profile_csv(locator("sales.csv"), sample_limit=1)

    assert profile.row_count == 2
    assert [column.name for column in profile.columns] == [
        "order_id",
        "customer",
        "amount",
    ]
    assert len(profile.sample_rows) == 1
    assert profile.grain == "one row per order_id"


def test_xlsx_sheet_name_is_part_of_the_dataset_identity(tmp_path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["record_id", "value"])
    first.append([1, 10])
    second = workbook.create_sheet("Second")
    second.append(["record_id", "value"])
    second.append([1, 20])
    workbook.save(tmp_path / "sheets.xlsx")
    workbook.close()
    profiler = profiler_for(tmp_path)

    first_profile = profiler.profile_xlsx(locator("sheets.xlsx"), sheet_name="First")
    second_profile = profiler.profile_xlsx(locator("sheets.xlsx"), sheet_name="Second")

    assert first_profile.relation_name == "First"
    assert second_profile.relation_name == "Second"
    assert first_profile.logical_id != second_profile.logical_id
    assert first_profile.version_id != second_profile.version_id


def test_sensitive_values_are_redacted_and_require_review(tmp_path: Path) -> None:
    write(
        tmp_path / "people.csv",
        "person_id,email,phone\n1,a@example.com,13800138000\n",
    )

    profile = profiler_for(tmp_path).profile_csv(locator("people.csv"), sample_limit=1)

    assert profile.sample_rows[0]["email"] == "[REDACTED]"
    assert profile.sample_rows[0]["phone"] == "[REDACTED]"
    assert {check.code for check in profile.evidence.checks} >= {
        "sensitive-column",
        "sample-redacted",
        "sensitive-sample",
    }
    assert profile.review_status == "needs-review"
    assert "a@example.com" not in profile.model_dump_json()
    assert "13800138000" not in profile.model_dump_json()


def test_inventory_is_recursive_and_records_metadata_without_opening_payloads(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    nested = tmp_path / "nested"
    nested.mkdir()
    payload = nested / "weights.pth"
    payload.write_bytes(b"must remain unopened")
    write(nested / "sales.csv", "record_id,value\n1,10\n")
    write(tmp_path / "legacy.xls", "legacy bytes are inventory only")
    write(tmp_path / "notes.txt", "not a dataset")
    real_open = Path.open

    def guarded_open(path: Path, *args: object, **kwargs: object) -> object:
        if path.resolve() == payload.resolve():
            raise AssertionError("quarantined payload was opened")
        return real_open(path, *args, **kwargs)

    monkeypatch.setattr(Path, "open", guarded_open)

    inventory = profiler_for(tmp_path).inventory_directory(locator("."))
    by_path = {item.relative_path: item for item in inventory}

    assert list(by_path) == [
        "legacy.xls",
        "nested/sales.csv",
        "nested/weights.pth",
        "notes.txt",
    ]
    assert by_path["nested/sales.csv"].extension == ".csv"
    assert by_path["nested/sales.csv"].byte_size > 0
    assert by_path["nested/sales.csv"].category == "tabular-dataset"
    assert by_path["legacy.xls"].disposition == "unsupported-deep-profile"
    assert by_path["nested/weights.pth"].disposition == "quarantined"
    assert by_path["notes.txt"].category == "unclassified"
    assert by_path["notes.txt"].disposition == "unclassified"


def test_csv_profile_computes_missingness_and_marks_unknown_grain_for_review(
    tmp_path: Path,
) -> None:
    write(tmp_path / "observations.csv", "label,amount\nA,10\nB,\n")

    profile = profiler_for(tmp_path).profile_csv(locator("observations.csv"))

    assert profile.missingness == {"label": 0.0, "amount": 0.5}
    amount = next(column for column in profile.columns if column.name == "amount")
    assert amount.missing_count == 1
    assert amount.missing_rate == 0.5
    assert amount.nullable is True
    assert profile.grain == "unknown"
    assert profile.review_status == "needs-review"
    grain_check = next(
        check for check in profile.evidence.checks if check.code == "grain-needs-review"
    )
    assert grain_check.details["blocking"] is False


def test_csv_profile_caps_samples_at_twenty_and_is_deterministic(tmp_path: Path) -> None:
    rows = "\n".join(f"{index},value-{index}" for index in range(1, 26))
    write(tmp_path / "records.csv", f"record_id,label\n{rows}\n")
    profiler = profiler_for(tmp_path)
    source = locator("records.csv")

    first = profiler.profile_csv(source, sample_limit=100)
    second = profiler.profile_csv(source, sample_limit=100)

    assert len(first.sample_rows) == 20
    assert first == second
    scope = next(
        check for check in first.evidence.checks if check.code == "statistics-scope"
    )
    assert scope.details["sample_limit"] == 20
    assert scope.details["missingness"] == "full-file"
    assert scope.details["uniqueness"] == "full-file"


def test_sensitive_value_pattern_redacts_an_opaque_column(tmp_path: Path) -> None:
    raw_value = "person@example.com"
    write(tmp_path / "contacts.csv", f"record_id,contact\n1,{raw_value}\n")

    profile = profiler_for(tmp_path).profile_csv(locator("contacts.csv"), sample_limit=1)

    assert profile.sample_rows[0]["contact"] == "[REDACTED]"
    contact = next(column for column in profile.columns if column.name == "contact")
    assert contact.sensitive_category == "email"
    assert raw_value not in profile.model_dump_json()


def test_sensitive_name_token_is_detected_inside_a_normalized_column(
    tmp_path: Path,
) -> None:
    raw_value = "private extension"
    write(
        tmp_path / "contacts.csv",
        f"record_id,primary_phone_number\n1,{raw_value}\n",
    )

    profile = profiler_for(tmp_path).profile_csv(locator("contacts.csv"))

    assert profile.sample_rows[0]["primary_phone_number"] == "[REDACTED]"
    assert raw_value not in profile.model_dump_json()


@pytest.mark.parametrize(
    ("column_name", "raw_value", "expected_category"),
    (
        ("homeAddress", "221B Baker Street", "address"),
        ("PrimaryPhoneNumber", "private extension", "phone"),
        ("EmailAddress", "private mailbox", "email"),
        ("idCardNo", "private identity token", "id-card"),
    ),
)
def test_camel_and_pascal_sensitive_names_never_persist_raw_samples(
    tmp_path: Path,
    column_name: str,
    raw_value: str,
    expected_category: str,
) -> None:
    write(
        tmp_path / "contacts.csv",
        f"record_id,{column_name}\n1,{raw_value}\n",
    )

    profile = profiler_for(tmp_path).profile_csv(locator("contacts.csv"))

    serialized = profile.model_dump_json()
    sensitive_column = next(
        column for column in profile.columns if column.name == column_name
    )
    check_codes = {check.code for check in profile.evidence.checks}
    assert profile.sample_rows[0][column_name] == "[REDACTED]"
    assert raw_value not in serialized
    assert sensitive_column.sensitive_category == expected_category
    assert check_codes >= {
        "sensitive-column",
        "sample-redacted",
        "sensitive-sample",
    }
    assert profile.review_status == "needs-review"


def test_profile_rejects_negative_sample_limits(tmp_path: Path) -> None:
    write(tmp_path / "records.csv", "record_id\n1\n")

    with pytest.raises(ValueError, match="negative"):
        profiler_for(tmp_path).profile_csv(locator("records.csv"), sample_limit=-1)


def test_xlsx_profile_uses_data_only_read_only_mode_and_redacts_samples(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook_path = tmp_path / "customers.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Customers"
    sheet.append(("customer_id", "email", "score", "formula_result"))
    sheet.append((1, "one@example.com", 10, "=1+1"))
    sheet.append((2, "two@example.com", None, "=2+2"))
    workbook.create_sheet("Lookup")
    workbook.save(workbook_path)
    real_load_workbook = dataset_profiler_module.load_workbook
    captured: dict[str, object] = {}

    def guarded_load_workbook(*args: object, **kwargs: object) -> object:
        captured.update(kwargs)
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(dataset_profiler_module, "load_workbook", guarded_load_workbook)

    profile = profiler_for(tmp_path).profile_xlsx(
        locator("customers.xlsx"),
        sample_limit=1,
    )

    assert profile.relation_name == "Customers"

    assert captured["read_only"] is True
    assert captured["data_only"] is True
    assert profile.sheets == ("Customers", "Lookup")
    assert profile.row_count == 2
    assert [column.name for column in profile.columns] == [
        "customer_id",
        "email",
        "score",
        "formula_result",
    ]
    assert profile.sample_rows[0]["email"] == "[REDACTED]"
    assert profile.sample_rows[0]["formula_result"] is None
    assert profile.missingness["score"] == 0.5
    assert profile.grain == "one row per customer_id"
    assert "one@example.com" not in profile.model_dump_json()
    assert "=1+1" not in profile.model_dump_json()


def test_xlsx_profile_can_select_a_sheet(tmp_path: Path) -> None:
    workbook_path = tmp_path / "multi.xlsx"
    workbook = Workbook()
    workbook.active.append(("ignored_id",))
    selected = workbook.create_sheet("Selected")
    selected.append(("row_id", "value"))
    selected.append((1, "kept"))
    workbook.save(workbook_path)

    profile = profiler_for(tmp_path).profile_xlsx(
        locator("multi.xlsx"),
        sheet_name="Selected",
    )

    assert profile.row_count == 1
    assert [column.name for column in profile.columns] == ["row_id", "value"]
    assert profile.sample_rows[0]["value"] == "kept"


def test_xlsx_header_names_are_globally_unique_without_losing_columns(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "colliding-headers.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("id", "id_2", "id"))
    sheet.append((1, None, 3))
    sheet.append((2, 20, None))
    workbook.save(workbook_path)

    profile = profiler_for(tmp_path).profile_xlsx(
        locator("colliding-headers.xlsx"),
        sample_limit=2,
    )

    names = [column.name for column in profile.columns]
    assert names == ["id", "id_2", "id_3"]
    assert len(names) == len(set(names)) == 3
    assert dict(profile.sample_rows[0]) == {"id": 1, "id_2": None, "id_3": 3}
    assert dict(profile.sample_rows[1]) == {"id": 2, "id_2": 20, "id_3": None}
    assert profile.missingness == {"id": 0.0, "id_2": 0.5, "id_3": 0.5}


def test_parquet_profile_is_bounded_redacted_and_byte_stable(tmp_path: Path) -> None:
    parquet_path = tmp_path / "customers.parquet"
    with duckdb.connect(database=":memory:") as connection:
        connection.execute(
            """
            CREATE TABLE customers AS
            SELECT
                value AS record_id,
                'private street ' || value::VARCHAR AS homeAddress,
                CASE WHEN value = 7 THEN NULL ELSE value * 1.5 END AS amount,
                CASE WHEN value % 2 = 0 THEN 'active' ELSE 'new' END AS segment
            FROM range(1, 26) AS rows(value)
            """
        )
        connection.table("customers").write_parquet(str(parquet_path))

    profiler = profiler_for(tmp_path)
    first = profiler.profile_parquet(locator("customers.parquet"), sample_limit=100)
    second = profiler.profile_parquet(locator("customers.parquet"), sample_limit=100)

    assert first == second
    assert first.format == "parquet"
    assert first.row_count == 25
    assert [column.name for column in first.columns] == [
        "record_id",
        "homeAddress",
        "amount",
        "segment",
    ]
    assert len(first.sample_rows) == 20
    assert first.missingness["amount"] == pytest.approx(1 / 25)
    assert all(row["homeAddress"] == "[REDACTED]" for row in first.sample_rows)
    serialized = first.model_dump_json()
    assert "private street" not in serialized
    assert first.review_status == "needs-review"
    assert profiler.inventory_directory(locator("."))[0].disposition == "profile-supported"


def test_legacy_xls_is_unsupported_without_opening_or_hashing(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    legacy_path = tmp_path / "legacy.xls"
    legacy_path.write_bytes(b"not a workbook parser input")
    real_open = Path.open

    def guarded_open(path: Path, *args: object, **kwargs: object) -> object:
        if path.resolve() == legacy_path.resolve():
            raise AssertionError("legacy XLS was opened")
        return real_open(path, *args, **kwargs)

    def guarded_hash(path: Path) -> str:
        if path.resolve() == legacy_path.resolve():
            raise AssertionError("legacy XLS was hashed")
        return "0" * 64

    monkeypatch.setattr(Path, "open", guarded_open)
    monkeypatch.setattr(dataset_profiler_module, "stream_sha256", guarded_hash)

    profiler = profiler_for(tmp_path)
    inventory = profiler.inventory_directory(locator("."))
    profile = profiler.profile_xlsx(locator("legacy.xls"))

    assert inventory[0].disposition == "unsupported-deep-profile"
    assert profile.format == "xls"
    assert profile.review_status == "unsupported"
    assert profile.row_count == 0
    assert profile.columns == ()
    check = next(
        check
        for check in profile.evidence.checks
        if check.code == "unsupported-deep-profile"
    )
    assert check.details["opened"] is False
    assert check.details["hashed"] is False


@pytest.mark.reference_demo
def test_allowlisted_csv_reference_has_expected_shape() -> None:
    profile = reference_profiler().profile_csv(
        SourceLocator(root_id="reference-demo", relative_path="dataset/1-train.csv")
    )

    assert len(profile.columns) == 12
    assert profile.row_count > 0


@pytest.mark.reference_demo
def test_allowlisted_xlsx_reference_has_sheet_and_schema() -> None:
    profile = reference_profiler().profile_xlsx(
        SourceLocator(
            root_id="reference-demo",
            relative_path="AIExcelData/ex-17-RFM.xlsx",
        )
    )

    assert profile.sheets
    assert profile.columns


@pytest.mark.reference_demo
def test_reference_model_payload_is_quarantined_without_opening_or_hashing(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    configured_root = os.environ.get("COURSE_REFERENCE_ROOT")
    if not configured_root:
        pytest.skip("COURSE_REFERENCE_ROOT is required for reference_demo tests")
    payload_relative_path = "AIExcelData/weights/sam_vit_h_4b8939.pth"
    payload = (Path(configured_root) / payload_relative_path).resolve(strict=True)
    real_open = Path.open

    def guarded_open(path: Path, *args: object, **kwargs: object) -> object:
        if path.resolve() == payload:
            raise AssertionError("reference model payload was opened")
        return real_open(path, *args, **kwargs)

    def guarded_hash(path: Path) -> str:
        if path.resolve() == payload:
            raise AssertionError("reference model payload was hashed")
        return "0" * 64

    monkeypatch.setattr(Path, "open", guarded_open)
    monkeypatch.setattr(dataset_profiler_module, "stream_sha256", guarded_hash)
    inventory = reference_profiler().inventory_directory(
        SourceLocator(root_id="reference-demo", relative_path="AIExcelData/weights")
    )

    item = next(
        item for item in inventory if item.relative_path == payload_relative_path
    )
    assert item.disposition == "quarantined"
