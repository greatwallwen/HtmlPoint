"""Stat-only dataset inventory and bounded profiles for registered sources."""

from __future__ import annotations

import hashlib
import json
import re
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from decimal import Decimal
from pathlib import Path, PurePosixPath
from typing import Any, Literal, TypeAlias, cast
from uuid import uuid5
from zipfile import BadZipFile, ZipFile

import duckdb
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field

from course_helper.domain.common import ActorRef, ImmutableJsonValue, SourceLocator
from course_helper.domain.evidence import EvidenceCheck, EvidenceObject
from course_helper.domain.sources import DatasetAssetVersion, DatasetColumn
from course_helper.source_roots import (
    COURSE_STUDIO_ID_NAMESPACE,
    SourceRootRegistry,
    SourceRootViolation,
    candidate_logical_id,
    candidate_version_id,
    stream_sha256,
)


class DatasetInventoryItem(BaseModel):
    """Immutable stat metadata for one recursively inventoried file."""

    model_config = ConfigDict(extra="forbid", frozen=True, strict=True)

    relative_path: str = Field(min_length=1)
    extension: str
    byte_size: int = Field(ge=0)
    modified_at: datetime
    category: Literal[
        "tabular-dataset",
        "analytical-database",
        "quarantined-payload",
        "unclassified",
    ]
    disposition: Literal[
        "profile-supported",
        "inventory-only",
        "unsupported-deep-profile",
        "quarantined",
        "unclassified",
    ]


DatasetInventory: TypeAlias = tuple[DatasetInventoryItem, ...]


_QUARANTINED_EXTENSIONS = frozenset({".pt", ".pth", ".tmp", ".whl"})
_PROFILE_SUPPORTED_EXTENSIONS = frozenset({".csv", ".parquet", ".xlsx"})
_MAX_SAMPLE_LIMIT = 20
_XLSX_DISTINCT_LIMIT = 20480
_CSV_RELATION_SQL = "read_csv_auto(?, SAMPLE_SIZE=20480, ALL_VARCHAR=false)"
_PARQUET_RELATION_SQL = "read_parquet(?)"
_CSV_VIEW_NAME = "_course_dataset_profile"
_CHART_RELATION_NAME = "_course_chart_relation"
_CHART_XLSX_TYPES = frozenset(
    {"BOOLEAN", "BIGINT", "DOUBLE", "TIMESTAMP", "DATE", "TIME", "VARCHAR"}
)
_PRODUCER = "course-helper/dataset-profiler"
_PRODUCER_VERSION = "1"
_ACTOR = ActorRef(actor_type="service", actor_id=_PRODUCER)
_SENSITIVE_NAMES = ("id_card", "email", "mobile", "phone", "ssn", "address")
_EMAIL_PATTERN = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
_MOBILE_PATTERN = re.compile(r"^(?:\+?86[- ]?)?1[3-9]\d{9}$")
_PHONE_PATTERN = re.compile(r"^(?:\+?\d{1,3}[- ]?)?(?:\d[- ]?){7,14}\d$")
_SSN_PATTERN = re.compile(r"^\d{3}-\d{2}-\d{4}$")
_ID_CARD_PATTERN = re.compile(r"^\d{17}[0-9Xx]$")
_ACRONYM_BOUNDARY_PATTERN = re.compile(r"([A-Z]+)([A-Z][a-z])")
_CAMEL_BOUNDARY_PATTERN = re.compile(r"([a-z0-9])([A-Z])")


@dataclass(frozen=True)
class VerifiedDatasetRelation:
    """Server-only bound DuckDB relation for one exact profiled dataset."""

    relation_sql: str
    parameters: tuple[str, ...]
    column_names: tuple[str, ...]
    column_types: tuple[str, ...]
    row_count: int


class DatasetProfiler:
    """Inventory registered roots without reading payload bytes."""

    def __init__(self, source_roots: SourceRootRegistry) -> None:
        self._source_roots = source_roots

    def inventory_directory(self, locator: SourceLocator) -> DatasetInventory:
        directory = self._source_roots.resolve_directory(locator)
        items: list[DatasetInventoryItem] = []
        for candidate in directory.rglob("*"):
            try:
                relative = PurePosixPath(
                    locator.relative_path,
                    candidate.relative_to(directory).as_posix(),
                ).as_posix()
                self._source_roots.resolve(
                    SourceLocator(
                        root_id=locator.root_id,
                        relative_path=relative,
                    )
                )
            except (OSError, RuntimeError, SourceRootViolation, ValueError):
                continue
            stat = candidate.stat()
            extension = candidate.suffix.lower()
            category, disposition = _classify_extension(extension)
            items.append(
                DatasetInventoryItem(
                    relative_path=relative,
                    extension=extension,
                    byte_size=stat.st_size,
                    modified_at=datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc),
                    category=category,
                    disposition=disposition,
                )
            )
        return tuple(
            sorted(
                items,
                key=lambda item: (item.relative_path.casefold(), item.relative_path),
            )
        )

    def profile_csv(
        self,
        locator: SourceLocator,
        sample_limit: int = _MAX_SAMPLE_LIMIT,
    ) -> DatasetAssetVersion:
        """Profile one allowlisted CSV with full aggregates and a bounded sample."""

        bounded_limit = _bounded_sample_limit(sample_limit)
        path = self._source_roots.resolve(locator)
        if path.suffix.lower() != ".csv":
            raise ValueError("profile_csv requires a .csv source")

        return self._profile_duckdb_relation(
            locator=locator,
            path=path,
            sample_limit=bounded_limit,
            format_name="csv",
            prepare_source=_prepare_csv_source,
        )

    def profile_parquet(
        self,
        locator: SourceLocator,
        sample_limit: int = _MAX_SAMPLE_LIMIT,
    ) -> DatasetAssetVersion:
        """Profile one allowlisted Parquet file without materializing it in Python."""

        bounded_limit = _bounded_sample_limit(sample_limit)
        path = self._source_roots.resolve(locator)
        if path.suffix.lower() != ".parquet":
            raise ValueError("profile_parquet requires a .parquet source")

        return self._profile_duckdb_relation(
            locator=locator,
            path=path,
            sample_limit=bounded_limit,
            format_name="parquet",
            prepare_source=_prepare_parquet_source,
        )

    def _profile_duckdb_relation(
        self,
        *,
        locator: SourceLocator,
        path: Path,
        sample_limit: int,
        format_name: Literal["csv", "parquet"],
        prepare_source: Callable[
            [duckdb.DuckDBPyConnection, Path],
            tuple[str, tuple[str, ...], Sequence[tuple[Any, ...]]],
        ],
    ) -> DatasetAssetVersion:
        """Share bounded DuckDB aggregates across CSV and Parquet sources."""

        content_digest = stream_sha256(path)
        source_time = _source_time(path)
        logical_id, version_id = _dataset_ids(locator, content_digest)
        config_digest = _profile_config_digest(
            format_name=format_name,
            sample_limit=sample_limit,
            sheet_name=None,
        )

        with duckdb.connect(database=":memory:") as connection:
            relation_sql, path_parameters, description = prepare_source(
                connection,
                path,
            )
            column_names = tuple(str(field[0]) for field in description)
            column_types = tuple(str(field[1]) for field in description)
            row_count = int(
                _execute_csv_query(
                    connection,
                    relation_sql,
                    path_parameters,
                    "count(*)",
                ).fetchone()[0]
            )
            missing_counts = _csv_missing_counts(
                connection,
                relation_sql,
                path_parameters,
                column_names,
            )
            raw_rows = tuple(
                _execute_csv_query(
                    connection,
                    relation_sql,
                    path_parameters,
                    "*",
                    suffix=" LIMIT ?",
                    extra_parameters=(sample_limit,),
                ).fetchall()
            )
            sensitive = _detect_sensitive_columns(column_names, raw_rows)
            distinct_counts, grain_column, grain_check = _csv_grain(
                connection=connection,
                relation_sql=relation_sql,
                path_parameters=path_parameters,
                column_names=column_names,
                missing_counts=missing_counts,
                sensitive=sensitive,
                row_count=row_count,
            )

        sample_rows = _redacted_sample_rows(column_names, raw_rows, sensitive)
        missingness = {
            name: (count / row_count if row_count else 0.0)
            for name, count in zip(column_names, missing_counts, strict=True)
        }
        columns = tuple(
            DatasetColumn(
                name=name,
                data_type=data_type,
                nullable=missing_count > 0,
                missing_count=missing_count,
                missing_rate=missingness[name],
                distinct_count=distinct_counts[index],
                sensitive_category=(
                    sensitive[index][0] if index in sensitive else None
                ),
            )
            for index, (name, data_type, missing_count) in enumerate(
                zip(column_names, column_types, missing_counts, strict=True)
            )
        )
        checks = _profile_checks(
            format_name=format_name,
            row_count=row_count,
            sample_count=len(sample_rows),
            sample_limit=sample_limit,
            statistics_scope="full-file",
            sensitive=sensitive,
            column_names=column_names,
            grain_check=grain_check,
        )
        grain = f"one row per {grain_column}" if grain_column is not None else "unknown"
        needs_review = bool(sensitive) or grain_column is None
        evidence = _evidence(
            locator=locator,
            version_id=version_id,
            config_digest=config_digest,
            source_time=source_time,
            row_count=row_count,
            column_count=len(columns),
            sample_count=len(sample_rows),
            statistics_scope="full-file",
            checks=checks,
            needs_review=needs_review,
        )
        return DatasetAssetVersion(
            logical_id=logical_id,
            version_id=version_id,
            revision=1,
            content_digest=content_digest,
            created_at=source_time,
            created_by=_ACTOR,
            locator=locator,
            format=format_name,
            row_count=row_count,
            columns=columns,
            grain=grain,
            missingness=missingness,
            sample_rows=sample_rows,
            review_status="needs-review" if needs_review else "ready",
            evidence=evidence,
        )

    def profile_xlsx(
        self,
        locator: SourceLocator,
        sheet_name: str | None = None,
        sample_limit: int = _MAX_SAMPLE_LIMIT,
    ) -> DatasetAssetVersion:
        """Profile one XLSX safely, or return stat-only evidence for legacy XLS."""

        bounded_limit = _bounded_sample_limit(sample_limit)
        path = self._source_roots.resolve(locator)
        extension = path.suffix.lower()
        if extension == ".xls":
            return _unsupported_xls_profile(
                locator=locator,
                path=path,
                sample_limit=bounded_limit,
                sheet_name=sheet_name,
            )
        if extension != ".xlsx":
            raise ValueError("profile_xlsx requires an .xlsx or .xls source")

        content_digest = stream_sha256(path)
        source_time = _source_time(path)
        config_digest = _profile_config_digest(
            format_name="xlsx",
            sample_limit=bounded_limit,
            sheet_name=sheet_name,
        )
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            sheets = tuple(workbook.sheetnames)
            if not sheets:
                raise ValueError("XLSX workbook has no worksheets")
            selected_sheet_name = sheet_name if sheet_name is not None else sheets[0]
            if selected_sheet_name not in sheets:
                raise ValueError("requested worksheet is not present")
            worksheet = workbook[selected_sheet_name]
            (
                column_names,
                column_types,
                row_count,
                missing_counts,
                raw_rows,
                distinct_counts,
                distinct_overflow,
                duplicate_candidates,
            ) = _scan_xlsx_worksheet(worksheet, bounded_limit)
        finally:
            workbook.close()

        logical_id, version_id = _dataset_ids(
            locator,
            content_digest,
            relation_name=selected_sheet_name,
        )

        sensitive = _detect_sensitive_columns(column_names, raw_rows)
        grain_column, grain_check, uniqueness_scope = _xlsx_grain(
            column_names=column_names,
            missing_counts=missing_counts,
            sensitive=sensitive,
            row_count=row_count,
            distinct_counts=distinct_counts,
            distinct_overflow=distinct_overflow,
            duplicate_candidates=duplicate_candidates,
        )
        sample_rows = _redacted_sample_rows(column_names, raw_rows, sensitive)
        missingness = {
            name: (count / row_count if row_count else 0.0)
            for name, count in zip(column_names, missing_counts, strict=True)
        }
        columns = tuple(
            DatasetColumn(
                name=name,
                data_type=data_type,
                nullable=missing_count > 0,
                missing_count=missing_count,
                missing_rate=missingness[name],
                distinct_count=(
                    distinct_counts.get(index)
                    if index not in distinct_overflow and index not in sensitive
                    else None
                ),
                sensitive_category=(
                    sensitive[index][0] if index in sensitive else None
                ),
            )
            for index, (name, data_type, missing_count) in enumerate(
                zip(column_names, column_types, missing_counts, strict=True)
            )
        )
        checks = _profile_checks(
            format_name="xlsx",
            row_count=row_count,
            sample_count=len(sample_rows),
            sample_limit=bounded_limit,
            statistics_scope=uniqueness_scope,
            sensitive=sensitive,
            column_names=column_names,
            grain_check=grain_check,
        )
        grain = f"one row per {grain_column}" if grain_column is not None else "unknown"
        needs_review = bool(sensitive) or grain_column is None
        evidence = _evidence(
            locator=locator,
            version_id=version_id,
            config_digest=config_digest,
            source_time=source_time,
            row_count=row_count,
            column_count=len(columns),
            sample_count=len(sample_rows),
            statistics_scope=uniqueness_scope,
            checks=checks,
            needs_review=needs_review,
        )
        return DatasetAssetVersion(
            logical_id=logical_id,
            version_id=version_id,
            revision=1,
            content_digest=content_digest,
            created_at=source_time,
            created_by=_ACTOR,
            locator=locator,
            format="xlsx",
            row_count=row_count,
            columns=columns,
            sheets=sheets,
            grain=grain,
            missingness=missingness,
            relation_name=selected_sheet_name,
            sample_rows=sample_rows,
            review_status="needs-review" if needs_review else "ready",
            evidence=evidence,
        )

    def prepare_verified_chart_relation(
        self,
        connection: duckdb.DuckDBPyConnection,
        dataset: DatasetAssetVersion,
        *,
        max_rows: int,
        max_bytes: int,
    ) -> VerifiedDatasetRelation:
        """Bind exact CSV/XLSX bytes to a bounded DuckDB relation for charts."""

        if (
            isinstance(max_rows, bool)
            or not isinstance(max_rows, int)
            or max_rows < 1
            or isinstance(max_bytes, bool)
            or not isinstance(max_bytes, int)
            or max_bytes < 1
        ):
            raise ValueError("chart source limits must be positive integers")
        if dataset.format not in {"csv", "xlsx"}:
            raise ValueError("chart relations currently require CSV or XLSX datasets")
        if dataset.row_count > max_rows:
            raise ValueError("dataset exceeds the chart row ceiling")
        path = self._source_roots.resolve(dataset.locator)
        if path.stat().st_size > max_bytes:
            raise ValueError("dataset exceeds the chart byte ceiling")
        expected_extension = ".csv" if dataset.format == "csv" else ".xlsx"
        governed_blob = (
            dataset.locator.root_id == "governed-upload"
            and path.suffix.lower() == ".blob"
        )
        if path.suffix.lower() != expected_extension and not governed_blob:
            raise ValueError("dataset format does not match its registered source")
        if stream_sha256(path) != dataset.content_digest:
            raise ValueError("dataset source digest changed")

        if dataset.format == "csv":
            relation_sql, parameters, description = _prepare_csv_source(connection, path)
            column_names = tuple(str(field[0]) for field in description)
            column_types = tuple(str(field[1]) for field in description)
            row_count = int(
                _execute_csv_query(
                    connection,
                    relation_sql,
                    parameters,
                    "count(*)",
                ).fetchone()[0]
            )
        else:
            relation_sql, parameters, column_names, column_types, row_count = (
                _prepare_xlsx_chart_relation(
                    connection,
                    path,
                    dataset=dataset,
                    max_rows=max_rows,
                )
            )

        expected_names = tuple(column.name for column in dataset.columns)
        expected_types = tuple(column.data_type for column in dataset.columns)
        if (
            column_names != expected_names
            or column_types != expected_types
            or row_count != dataset.row_count
            or row_count > max_rows
        ):
            raise ValueError("dataset schema or row count drifted")
        return VerifiedDatasetRelation(
            relation_sql=relation_sql,
            parameters=parameters,
            column_names=column_names,
            column_types=column_types,
            row_count=row_count,
        )

    def verify_dataset_source_digest(self, dataset: DatasetAssetVersion) -> bool:
        """Recheck exact source bytes after a chart query without exposing a path."""

        path = self._source_roots.resolve(dataset.locator)
        if stream_sha256(path) != dataset.content_digest:
            raise ValueError("dataset source digest changed")
        return True

def _classify_extension(
    extension: str,
) -> tuple[
    Literal[
        "tabular-dataset",
        "analytical-database",
        "quarantined-payload",
        "unclassified",
    ],
    Literal[
        "profile-supported",
        "inventory-only",
        "unsupported-deep-profile",
        "quarantined",
        "unclassified",
    ],
]:
    if extension in _QUARANTINED_EXTENSIONS:
        return "quarantined-payload", "quarantined"
    if extension in _PROFILE_SUPPORTED_EXTENSIONS:
        return "tabular-dataset", "profile-supported"
    if extension == ".xls":
        return "tabular-dataset", "unsupported-deep-profile"
    if extension == ".duckdb":
        return "analytical-database", "inventory-only"
    return "unclassified", "unclassified"


def _bounded_sample_limit(sample_limit: int) -> int:
    if isinstance(sample_limit, bool) or not isinstance(sample_limit, int):
        raise TypeError("sample_limit must be an integer")
    if sample_limit < 0:
        raise ValueError("sample_limit cannot be negative")
    return min(sample_limit, _MAX_SAMPLE_LIMIT)


def _prepare_csv_source(
    connection: duckdb.DuckDBPyConnection,
    path: Path,
) -> tuple[str, tuple[str, ...], Sequence[tuple[Any, ...]]]:
    try:
        cursor = connection.execute(
            f"SELECT * FROM {_CSV_RELATION_SQL} LIMIT ?",
            [str(path), 0],
        )
        return _CSV_RELATION_SQL, (str(path),), tuple(cursor.description)
    except duckdb.Error:
        relation = connection.read_csv(
            str(path),
            sample_size=20480,
            all_varchar=False,
        )
        relation.create_view(_CSV_VIEW_NAME, replace=True)
        cursor = connection.execute(f'SELECT * FROM "{_CSV_VIEW_NAME}" LIMIT ?', [0])
        return f'"{_CSV_VIEW_NAME}"', (), tuple(cursor.description)


def _prepare_parquet_source(
    connection: duckdb.DuckDBPyConnection,
    path: Path,
) -> tuple[str, tuple[str, ...], Sequence[tuple[Any, ...]]]:
    cursor = connection.execute(
        f"SELECT * FROM {_PARQUET_RELATION_SQL} LIMIT ?",
        [str(path), 0],
    )
    return _PARQUET_RELATION_SQL, (str(path),), tuple(cursor.description)


def _prepare_xlsx_chart_relation(
    connection: duckdb.DuckDBPyConnection,
    path: Path,
    *,
    dataset: DatasetAssetVersion,
    max_rows: int,
) -> tuple[str, tuple[str, ...], tuple[str, ...], tuple[str, ...], int]:
    _validate_xlsx_chart_package(path)
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        if not workbook.sheetnames:
            raise ValueError("XLSX workbook has no worksheets")
        if dataset.relation_name is None:
            raise ValueError("registered XLSX relation is not pinned")
        selected = dataset.relation_name
        if selected not in workbook.sheetnames:
            raise ValueError("registered XLSX worksheet is not present")
        rows = workbook[selected].iter_rows(values_only=True)
        header = next(rows, None)
        column_names = _xlsx_column_names(tuple(header or ()))
        column_types = tuple(column.data_type for column in dataset.columns)
        observed_types: list[set[str]] = [set() for _ in column_names]
        values: list[tuple[Any, ...]] = []
        for source_row in rows:
            normalized = tuple(source_row[: len(column_names)]) + (None,) * max(
                0, len(column_names) - len(source_row)
            )
            if not any(value is not None for value in normalized):
                continue
            for index, value in enumerate(normalized):
                if value is not None:
                    observed_types[index].add(_xlsx_value_type(value))
            values.append(normalized)
            if len(values) > max_rows:
                raise ValueError("dataset exceeds the chart row ceiling")
    finally:
        workbook.close()

    if (
        not column_names
        or len(column_names) != len(column_types)
        or any(data_type not in _CHART_XLSX_TYPES for data_type in column_types)
        or tuple(_merged_xlsx_type(types) for types in observed_types) != column_types
        or _detect_sensitive_columns(column_names, tuple(values))
    ):
        raise ValueError("registered XLSX schema is invalid")
    connection.execute(f'DROP TABLE IF EXISTS "{_CHART_RELATION_NAME}"')
    definitions = ", ".join(
        f"{_quote_identifier(name)} {data_type}"
        for name, data_type in zip(column_names, column_types, strict=True)
    )
    connection.execute(f'CREATE TEMP TABLE "{_CHART_RELATION_NAME}" ({definitions})')
    if values:
        placeholders = ", ".join("?" for _ in column_names)
        try:
            connection.executemany(
                f'INSERT INTO "{_CHART_RELATION_NAME}" VALUES ({placeholders})',
                values,
            )
        except duckdb.Error as error:
            raise ValueError("XLSX values do not match the registered schema") from error
    return (
        f'"{_CHART_RELATION_NAME}"',
        (),
        column_names,
        column_types,
        len(values),
    )


def _validate_xlsx_chart_package(path: Path) -> None:
    """Reject ambiguous or expansion-heavy packages before openpyxl reads XML."""

    try:
        with ZipFile(path, "r") as archive:
            infos = tuple(archive.infolist())
            if len(infos) > 10_000:
                raise ValueError("XLSX package contains too many members")
            names: set[str] = set()
            uncompressed_bytes = 0
            for info in infos:
                member = PurePosixPath(info.filename)
                if (
                    not info.filename
                    or "\\" in info.filename
                    or member.is_absolute()
                    or ".." in member.parts
                    or info.filename in names
                    or info.flag_bits & 0x1
                    or info.file_size > 64 * 1024 * 1024
                ):
                    raise ValueError("XLSX package member is unsafe or ambiguous")
                names.add(info.filename)
                uncompressed_bytes += info.file_size
                if uncompressed_bytes > 128 * 1024 * 1024:
                    raise ValueError("XLSX package expansion exceeds the chart ceiling")
    except (BadZipFile, OSError) as error:
        raise ValueError("XLSX package is invalid") from error


def _execute_csv_query(
    connection: duckdb.DuckDBPyConnection,
    relation_sql: str,
    path_parameters: tuple[str, ...],
    projection: str,
    *,
    suffix: str = "",
    extra_parameters: tuple[int, ...] = (),
) -> duckdb.DuckDBPyConnection:
    query = f"SELECT {projection} FROM {relation_sql}{suffix}"
    return connection.execute(query, [*path_parameters, *extra_parameters])


def _csv_missing_counts(
    connection: duckdb.DuckDBPyConnection,
    relation_sql: str,
    path_parameters: tuple[str, ...],
    column_names: tuple[str, ...],
) -> tuple[int, ...]:
    projection = ", ".join(
        f"count(*) FILTER (WHERE {_quote_identifier(name)} IS NULL)"
        for name in column_names
    )
    if not projection:
        return ()
    values = _execute_csv_query(
        connection,
        relation_sql,
        path_parameters,
        projection,
    ).fetchone()
    return tuple(int(value) for value in values)


def _csv_grain(
    *,
    connection: duckdb.DuckDBPyConnection,
    relation_sql: str,
    path_parameters: tuple[str, ...],
    column_names: tuple[str, ...],
    missing_counts: tuple[int, ...],
    sensitive: Mapping[int, tuple[str, int]],
    row_count: int,
) -> tuple[tuple[int | None, ...], str | None, EvidenceCheck]:
    distinct_counts: list[int | None] = [None] * len(column_names)
    grain_column: str | None = None
    for index, (name, missing_count) in enumerate(
        zip(column_names, missing_counts, strict=True)
    ):
        if index in sensitive or missing_count or not _is_id_like(name):
            continue
        distinct_count = int(
            _execute_csv_query(
                connection,
                relation_sql,
                path_parameters,
                f"count(DISTINCT {_quote_identifier(name)})",
            ).fetchone()[0]
        )
        distinct_counts[index] = distinct_count
        if row_count > 0 and grain_column is None and distinct_count == row_count:
            grain_column = name

    if grain_column is None:
        check = EvidenceCheck(
            code="grain-needs-review",
            status="warning",
            message="No non-sensitive ID-like column was verified as unique",
            details={"blocking": False, "statistics_scope": "full-file"},
        )
    else:
        distinct_count = distinct_counts[column_names.index(grain_column)]
        check = EvidenceCheck(
            code="grain-verified",
            status="passed",
            message="A non-sensitive ID-like column is unique and non-null",
            details={
                "column": grain_column,
                "row_count": row_count,
                "distinct_count": distinct_count,
                "aggregate_query": (
                    f"SELECT count(DISTINCT {_quote_identifier(grain_column)})"
                ),
                "confidence": "verified-full-file",
            },
        )
    return tuple(distinct_counts), grain_column, check


def _scan_xlsx_worksheet(
    worksheet: Any,
    sample_limit: int,
) -> tuple[
    tuple[str, ...],
    tuple[str, ...],
    int,
    tuple[int, ...],
    tuple[tuple[Any, ...], ...],
    dict[int, int],
    frozenset[int],
    frozenset[int],
]:
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return (), (), 0, (), (), {}, frozenset(), frozenset()
    column_names = _xlsx_column_names(tuple(header))
    column_count = len(column_names)
    missing_counts = [0] * column_count
    type_names: list[set[str]] = [set() for _ in range(column_count)]
    raw_rows: list[tuple[Any, ...]] = []
    candidate_indices = tuple(
        index
        for index, name in enumerate(column_names)
        if _is_id_like(name) and _sensitive_name_category(name) is None
    )
    distinct_values: dict[int, set[tuple[str, str]]] = {
        index: set() for index in candidate_indices
    }
    distinct_overflow: set[int] = set()
    duplicate_candidates: set[int] = set()
    row_count = 0

    for source_row in rows:
        normalized_row = tuple(source_row[:column_count]) + (None,) * max(
            0,
            column_count - len(source_row),
        )
        if not any(value is not None for value in normalized_row):
            continue
        row_count += 1
        if len(raw_rows) < sample_limit:
            raw_rows.append(normalized_row)
        for index, value in enumerate(normalized_row):
            if value is None:
                missing_counts[index] += 1
            else:
                type_names[index].add(_xlsx_value_type(value))
        for index in candidate_indices:
            value = normalized_row[index]
            if value is None or index in distinct_overflow:
                continue
            marker = (type(value).__name__, str(value))
            seen = distinct_values[index]
            if marker in seen:
                duplicate_candidates.add(index)
            elif len(seen) >= _XLSX_DISTINCT_LIMIT:
                distinct_overflow.add(index)
            else:
                seen.add(marker)

    distinct_counts = {
        index: len(values) for index, values in distinct_values.items()
    }
    return (
        column_names,
        tuple(_merged_xlsx_type(names) for names in type_names),
        row_count,
        tuple(missing_counts),
        tuple(raw_rows),
        distinct_counts,
        frozenset(distinct_overflow),
        frozenset(duplicate_candidates),
    )


def _xlsx_grain(
    *,
    column_names: tuple[str, ...],
    missing_counts: tuple[int, ...],
    sensitive: Mapping[int, tuple[str, int]],
    row_count: int,
    distinct_counts: Mapping[int, int],
    distinct_overflow: frozenset[int],
    duplicate_candidates: frozenset[int],
) -> tuple[str | None, EvidenceCheck, str]:
    uniqueness_scope = "sampled" if distinct_overflow else "full-file"
    grain_column: str | None = None
    for index, (name, missing_count) in enumerate(
        zip(column_names, missing_counts, strict=True)
    ):
        if (
            index in sensitive
            or index in distinct_overflow
            or index in duplicate_candidates
            or missing_count
            or not _is_id_like(name)
        ):
            continue
        if row_count > 0 and distinct_counts.get(index) == row_count:
            grain_column = name
            break

    if grain_column is None:
        check = EvidenceCheck(
            code="grain-needs-review",
            status="warning",
            message="No non-sensitive ID-like column was verified as unique",
            details={"blocking": False, "statistics_scope": uniqueness_scope},
        )
    else:
        index = column_names.index(grain_column)
        check = EvidenceCheck(
            code="grain-verified",
            status="passed",
            message="A non-sensitive ID-like column is unique and non-null",
            details={
                "column": grain_column,
                "row_count": row_count,
                "distinct_count": distinct_counts[index],
                "aggregate_query": "bounded in-memory distinct set",
                "confidence": "verified-full-file",
            },
        )
    return grain_column, check, uniqueness_scope


def _detect_sensitive_columns(
    column_names: tuple[str, ...],
    rows: tuple[tuple[Any, ...], ...],
) -> dict[int, tuple[str, int]]:
    detected: dict[int, tuple[str, int]] = {}
    for index, name in enumerate(column_names):
        name_category = _sensitive_name_category(name)
        values = tuple(row[index] for row in rows)
        if name_category is not None:
            detected[index] = (
                name_category,
                sum(value is not None for value in values),
            )
            continue
        categories = tuple(
            category
            for value in values
            if value is not None
            for category in (_sensitive_value_category(value),)
            if category is not None
        )
        if categories:
            category = categories[0]
            detected[index] = (category, categories.count(category))
    return detected


def _sensitive_name_category(name: str) -> str | None:
    acronym_split = _ACRONYM_BOUNDARY_PATTERN.sub(r"\1_\2", name)
    camel_split = _CAMEL_BOUNDARY_PATTERN.sub(r"\1_\2", acronym_split)
    normalized = re.sub(r"[^a-z0-9]+", "_", camel_split.casefold()).strip("_")
    tokenized = f"_{normalized}_"
    for sensitive_name in _SENSITIVE_NAMES:
        if f"_{sensitive_name}_" in tokenized:
            return sensitive_name.replace("_", "-")
    return None


def _sensitive_value_category(value: Any) -> str | None:
    text = str(value).strip()
    if _EMAIL_PATTERN.fullmatch(text):
        return "email"
    if _SSN_PATTERN.fullmatch(text):
        return "ssn"
    if _ID_CARD_PATTERN.fullmatch(text):
        return "id-card"
    if _MOBILE_PATTERN.fullmatch(text):
        return "mobile"
    if _PHONE_PATTERN.fullmatch(text):
        return "phone"
    return None


def _xlsx_column_names(header: tuple[Any, ...]) -> tuple[str, ...]:
    names: list[str] = []
    used_names: set[str] = set()
    for index, value in enumerate(header, start=1):
        base = str(value).strip() if value is not None else ""
        if not base:
            base = f"column_{index}"
        candidate = base
        suffix = 2
        while candidate in used_names:
            candidate = f"{base}_{suffix}"
            suffix += 1
        used_names.add(candidate)
        names.append(candidate)
    return tuple(names)


def _xlsx_value_type(value: Any) -> str:
    if isinstance(value, bool):
        return "BOOLEAN"
    if isinstance(value, int):
        return "BIGINT"
    if isinstance(value, (float, Decimal)):
        return "DOUBLE"
    if isinstance(value, datetime):
        return "TIMESTAMP"
    if isinstance(value, date):
        return "DATE"
    if isinstance(value, time):
        return "TIME"
    return "VARCHAR"


def _merged_xlsx_type(types: set[str]) -> str:
    if not types:
        return "VARCHAR"
    if len(types) == 1:
        return next(iter(types))
    if types <= {"BIGINT", "DOUBLE"}:
        return "DOUBLE"
    return "VARCHAR"


def _redacted_sample_rows(
    column_names: tuple[str, ...],
    rows: tuple[tuple[Any, ...], ...],
    sensitive: Mapping[int, tuple[str, int]],
) -> tuple[Mapping[str, ImmutableJsonValue], ...]:
    return tuple(
        {
            name: (
                "[REDACTED]" if index in sensitive else _json_value(row[index])
            )
            for index, name in enumerate(column_names)
        }
        for row in rows
    )


def _json_value(value: Any) -> ImmutableJsonValue:
    if value is None or isinstance(value, (str, bool, int, float)):
        return cast(ImmutableJsonValue, value)
    if isinstance(value, Decimal):
        return cast(ImmutableJsonValue, str(value))
    if isinstance(value, (datetime, date, time)):
        return cast(ImmutableJsonValue, value.isoformat())
    return cast(ImmutableJsonValue, str(value))


def _profile_checks(
    *,
    format_name: str,
    row_count: int,
    sample_count: int,
    sample_limit: int,
    statistics_scope: str,
    sensitive: Mapping[int, tuple[str, int]],
    column_names: tuple[str, ...],
    grain_check: EvidenceCheck,
) -> tuple[EvidenceCheck, ...]:
    checks: list[EvidenceCheck] = [
        EvidenceCheck(
            code="statistics-scope",
            status="passed",
            message="Dataset statistics and sample scope were recorded",
            details={
                "format": format_name,
                "row_count": "full-file",
                "missingness": "full-file",
                "uniqueness": statistics_scope,
                "sample": "bounded",
                "sample_limit": sample_limit,
                "sample_count": sample_count,
            },
        ),
        grain_check,
    ]
    for index, (category, detection_count) in sensitive.items():
        checks.append(
            EvidenceCheck(
                code="sensitive-column",
                status="warning",
                message="A sampled column was classified as sensitive",
                details={
                    "column": column_names[index],
                    "category": category,
                    "detection_count": detection_count,
                },
            )
        )
    if sensitive:
        checks.extend(
            (
                EvidenceCheck(
                    code="sample-redacted",
                    status="passed",
                    message="Every sampled value in sensitive columns was redacted",
                    details={
                        "column_count": len(sensitive),
                        "redacted_value_count": len(sensitive) * sample_count,
                    },
                ),
                EvidenceCheck(
                    code="sensitive-sample",
                    status="failed",
                    message="Sensitive columns require blocking human review",
                    details={"blocking": True, "column_count": len(sensitive)},
                ),
            )
        )
    return tuple(checks)


def _evidence(
    *,
    locator: SourceLocator,
    version_id: str,
    config_digest: str,
    source_time: datetime,
    row_count: int,
    column_count: int,
    sample_count: int,
    statistics_scope: str,
    checks: tuple[EvidenceCheck, ...],
    needs_review: bool,
) -> EvidenceObject:
    has_failed_check = any(check.status == "failed" for check in checks)
    return EvidenceObject(
        evidence_id=str(
            uuid5(
                COURSE_STUDIO_ID_NAMESPACE,
                f"evidence\0dataset-profile\0{version_id}\0{config_digest}",
            )
        ),
        kind="dataset-profile",
        subject_version_id=version_id,
        status=(
            "degraded" if has_failed_check else "warning" if needs_review else "verified"
        ),
        input_summary={
            "source_locator": locator.model_dump(mode="json"),
            "profile_config_digest": config_digest,
        },
        output_summary={
            "row_count": row_count,
            "column_count": column_count,
            "sample_count": sample_count,
            "statistics_scope": statistics_scope,
        },
        producer=_PRODUCER,
        producer_version=_PRODUCER_VERSION,
        started_at=source_time,
        finished_at=source_time,
        duration_ms=0,
        checks=checks,
    )


def _dataset_ids(
    locator: SourceLocator,
    content_digest: str,
    *,
    relation_name: str | None = None,
) -> tuple[str, str]:
    semantic_locator = f"{locator.root_id}:{locator.relative_path}"
    if relation_name is not None:
        semantic_locator = json.dumps(
            {
                "root_id": locator.root_id,
                "relative_path": locator.relative_path,
                "relation_name": relation_name,
            },
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
    logical_id = candidate_logical_id(
        "dataset",
        semantic_locator,
    )
    return logical_id, candidate_version_id(logical_id, (), content_digest)


def _unsupported_xls_profile(
    *,
    locator: SourceLocator,
    path: Path,
    sample_limit: int,
    sheet_name: str | None,
) -> DatasetAssetVersion:
    stat = path.stat()
    content_digest = hashlib.sha256(
        json.dumps(
            {
                "root_id": locator.root_id,
                "relative_path": locator.relative_path,
                "byte_size": stat.st_size,
                "modified_ns": stat.st_mtime_ns,
            },
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()
    source_time = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc)
    logical_id, version_id = _dataset_ids(locator, content_digest)
    config_digest = _profile_config_digest(
        format_name="xls",
        sample_limit=sample_limit,
        sheet_name=sheet_name,
    )
    checks = (
        EvidenceCheck(
            code="unsupported-deep-profile",
            status="skipped",
            message="Legacy XLS is inventoried but never opened by this profiler",
            details={
                "extension": ".xls",
                "opened": False,
                "hashed": False,
                "statistics_scope": "unsupported",
            },
        ),
    )
    evidence = _evidence(
        locator=locator,
        version_id=version_id,
        config_digest=config_digest,
        source_time=source_time,
        row_count=0,
        column_count=0,
        sample_count=0,
        statistics_scope="unsupported",
        checks=checks,
        needs_review=True,
    )
    return DatasetAssetVersion(
        logical_id=logical_id,
        version_id=version_id,
        revision=1,
        content_digest=content_digest,
        created_at=source_time,
        created_by=_ACTOR,
        locator=locator,
        format="xls",
        row_count=0,
        columns=(),
        grain="unknown",
        review_status="unsupported",
        evidence=evidence,
    )


def _profile_config_digest(
    *,
    format_name: str,
    sample_limit: int,
    sheet_name: str | None,
) -> str:
    payload = json.dumps(
        {
            "format": format_name,
            "producer_version": _PRODUCER_VERSION,
            "sample_limit": sample_limit,
            "sheet_name": sheet_name,
        },
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _source_time(path: Path) -> datetime:
    return datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc)


def _quote_identifier(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _is_id_like(name: str) -> bool:
    normalized = re.sub(r"[^a-z0-9]+", "_", name.casefold()).strip("_")
    return normalized == "id" or normalized.endswith("_id") or normalized.startswith("id_")
